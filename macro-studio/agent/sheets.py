"""Reading a column out of a spreadsheet, without opening Excel.

The obvious way to get a list of order numbers out of a workbook is to
open it, select the column, Ctrl+C and read the clipboard. It works right
up until anything else wants the foreground: Excel ignores posted
keystrokes, so those steps need real ones, which means the workbook has
to stay focused for the whole run -- and a run that copies a number,
searches for it, copies the next one, and repeats a hundred times is a
long time to ask someone not to touch their own machine.

Reading the file directly costs none of that. The rest of the macro is
CDP web steps, which never need focus either, so the whole thing runs in
the background while the user keeps working -- which was the point of the
tool.

Values come back as strings, formatted the way they read on screen rather
than the way they are stored: an order number is 1001, not 1001.0, and a
search for "1001.0" finds nothing.
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import locale
from pathlib import Path

MAX_ROWS = 5000

# The key a row keeps its own one-line form under. Not a column name: no
# spreadsheet heading is a bare "#", and no variable is written {{row_#}}.
ROW_TEXT = "#"


def _clean(value) -> str:
    """One cell, as text a human would recognise it by."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _column_index(column: str, header: list[str]) -> int:
    """Column letters and header names both land here, because both are
    what people actually have: a column letter is what Excel shows in the
    grid, and a header name is what survives someone inserting a column
    in front of it.

    The heading is tried first, and this is the whole reason the function
    exists: plenty of real headings -- SKU, ID, QTY -- are also valid
    column letters, and reading SKU as column 12,871 finds nothing while
    looking exactly like an empty spreadsheet."""
    column = (column or "A").strip() or "A"
    want = column.lower()
    for i, cell in enumerate(header):
        if cell.strip().lower() == want:
            return i
    if column.isalpha() and len(column) <= 3:
        index = 0
        for char in column.upper():
            index = index * 26 + (ord(char) - ord("A") + 1)
        return index - 1
    named = ", ".join(c for c in header if c)[:120]
    raise RuntimeError(f'No column called "{column}" -- the heading row has: {named or "(nothing)"}.')


def _delimiter(sample: str) -> str:
    """Whichever separator the first real line has most of.

    csv.Sniffer is the library answer and it gives up on short files -- a
    two-column export with four rows is exactly the size it fails on, and
    its failure mode is silent: every row comes back as one long cell,
    which then reads as a spreadsheet with nothing in it. A machine set to
    Lithuanian or German locale writes semicolons, so this is not an edge
    case here."""
    for line in sample.splitlines():
        if not line.strip():
            continue
        counts = {sep: line.count(sep) for sep in (",", ";", "\t", "|")}
        best = max(counts, key=lambda sep: counts[sep])
        return best if counts[best] else ","
    return ","


def _looks_like_cjk_mojibake(text: str) -> bool:
    """Three accented Latin letters in a row is the signature of Chinese
    read as Western text: 蓝色杯子 decoded as cp1252 comes out À¶É«±­×Ó.
    Real Western text almost never stacks three of them, and Lithuanian's
    own letters live above U+0100, out of this range entirely."""
    run = 0
    for char in text[:4000]:
        if "" <= char <= "ÿ":
            run += 1
            if run >= 3:
                return True
        else:
            run = 0
    return False


def _decode_csv(path: Path, encoding: str) -> str:
    """A CSV carries no record of how it was written, and Excel writes it
    in whatever the machine's ANSI codepage is -- cp1252 here, cp936 on a
    Chinese Windows. Guessing wrong doesn't fail, it silently produces
    text nobody can search for, so the guess is checked for the one
    pattern that says it went wrong and the field can override it."""
    raw = path.read_bytes()
    if encoding:
        try:
            return raw.decode(encoding, errors="strict")
        except (UnicodeDecodeError, LookupError) as exc:
            raise RuntimeError(f'"{path.name}" is not {encoding} text: {exc}')
    try:
        return raw.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError:
        pass
    ansi = locale.getpreferredencoding(False) or "cp1252"
    for candidate in (ansi, "cp1252"):
        try:
            text = raw.decode(candidate, errors="strict")
        except (UnicodeDecodeError, LookupError):
            continue
        if _looks_like_cjk_mojibake(text):
            try:
                return raw.decode("gb18030", errors="strict")
            except UnicodeDecodeError:
                pass
        return text
    return raw.decode("cp1252", errors="replace")


def _rows_from_csv(path: Path, limit: int, encoding: str = "") -> list[list[str]]:
    text = _decode_csv(path, encoding)
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=_delimiter(text[:4096]))
    return [[_clean(cell) for cell in row] for row in reader][:limit]


def _rows_from_workbook(path: Path, sheet: str, limit: int) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover -- only if the venv is half-built
        raise RuntimeError("The spreadsheet reader needs openpyxl -- run start.bat once to install it.")
    try:
        # data_only: a cell holding =VLOOKUP(...) has both a formula and
        # the value Excel last calculated, and the value is the only half
        # anyone wants to search for. A workbook that has never been
        # opened and saved has no cached values at all, which is why the
        # step says so rather than returning a column of blanks.
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f'Could not open "{path.name}": {exc}')
    try:
        if sheet:
            wanted = sheet.strip().lower()
            match = next((name for name in book.sheetnames if name.strip().lower() == wanted), None)
            if match is None:
                raise RuntimeError(f'That workbook has no sheet called "{sheet}" -- it has: '
                                   + ", ".join(book.sheetnames))
            worksheet = book[match]
        else:
            worksheet = book[book.sheetnames[0]]
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([_clean(cell) for cell in row])
            if len(rows) >= limit:
                break
        return rows
    finally:
        book.close()


def _index_to_letter(index: int) -> str:
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _load(path: str, sheet: str, first_row: int, limit: int,
          encoding: str) -> tuple[list[list[str]], list[str], int]:
    """The rows, the heading row, and the capped limit -- everything both
    readers need before they differ."""
    target = Path(path)
    if not target.is_file():
        raise RuntimeError(f'There is no file at "{path}".')
    limit = min(max(1, limit), MAX_ROWS)
    first_row = max(1, first_row)
    # The limit counts values wanted, but the rows read have to include
    # the ones skipped above the data.
    wanted = first_row - 1 + limit
    rows = (_rows_from_csv(target, wanted, encoding)
            if target.suffix.lower() == ".csv"
            else _rows_from_workbook(target, sheet, wanted))
    if not rows:
        raise RuntimeError(f'"{target.name}" has no rows in it.')
    # The headings are the row directly above where the data starts, not
    # necessarily row 1 -- exports padded with a title and a blank line
    # are common enough to matter.
    return rows, rows[max(0, first_row - 2)], limit


def _parse_columns(spec: str, header: list[str]) -> list[tuple[str, int]]:
    """"A-G", "A,C,F", "SKU, Qty", or blank for everything the heading row
    has. Ranges are the reason this exists: a row someone works with is a
    span of columns, and writing out seven of them by hand is the kind of
    thing that gets one wrong."""
    spec = (spec or "").strip()
    if not spec:
        return [(_index_to_letter(i), i) for i in range(len(header))] or [("A", 0)]
    picked: list[tuple[str, int]] = []
    for part in spec.replace(":", "-").split(","):
        part = part.strip()
        if not part:
            continue
        halves = [half.strip() for half in part.split("-")] if "-" in part else [part]
        if len(halves) == 2 and all(h.isalpha() and len(h) <= 3 for h in halves):
            start, end = (_column_index(halves[0], []), _column_index(halves[1], []))
            if end < start:
                start, end = end, start
            picked.extend((_index_to_letter(i), i) for i in range(start, end + 1))
            continue
        index = _column_index(part, header)
        picked.append((_index_to_letter(index), index))
    # Same column twice (an overlapping range, a name that is also in a
    # range) would otherwise read it twice into the same key.
    seen, unique = set(), []
    for letter, index in picked:
        if index not in seen:
            seen.add(index)
            unique.append((letter, index))
    return unique


def read_rows(path: str, columns: str = "", sheet: str = "", first_row: int = 1,
              limit: int = 500, encoding: str = "") -> list[dict]:
    """Several columns at once: one dict per row, keyed by column letter
    and, when there is a heading row above the data, by heading too.

    Both keys are kept because both get used: `{{row_A}}` is what someone
    writes while looking at the spreadsheet, and `{{row_SKU}}` is what
    survives a column being inserted in front of it.

    Rows where every picked cell is blank are dropped -- that is the end
    of the data, or the gap between two blocks of it.

    Each row also carries the whole line under the key ROW_TEXT. A heading
    like SKU or ID is itself a valid column letter, so nothing downstream
    can tell the two kinds of key apart by looking -- and printing a row
    would show every cell twice. Writing the line once, here, where the
    order is still known, is the only place that stays correct."""
    rows, header, limit = _load(path, sheet, first_row, limit, encoding)
    picks = _parse_columns(columns, header)
    named = first_row >= 2  # no heading row above row 1 to take names from
    out: list[dict] = []
    for row in rows[max(0, first_row - 1):]:
        cells: dict[str, str] = {}
        for letter, index in picks:
            value = row[index] if index < len(row) else ""
            cells[letter] = value
            heading = header[index].strip() if named and index < len(header) else ""
            # A heading with a brace or a colon in it can't be written as
            # {{row_name}}, so it stays reachable by its letter only.
            if heading and heading not in cells and not any(c in heading for c in "{}:"):
                cells[heading] = value
        if any(cells.values()):
            cells[ROW_TEXT] = " | ".join(cells[letter] for letter, _ in picks if cells[letter])
            out.append(cells)
        if len(out) >= limit:
            break
    return out


def read_column(path: str, column: str = "A", sheet: str = "", first_row: int = 1,
                limit: int = 500, encoding: str = "") -> list[str]:
    """The values in one column, top to bottom, blanks dropped.

    Blanks are dropped rather than kept as empty strings: they are the
    gap between two blocks of data or the ragged end of the sheet, and a
    macro that searched a site for "" once per empty row would be doing
    something nobody asked for. `first_row` is the row the data starts
    on, counting the way Excel counts -- 2 when row 1 is the header."""
    rows, header, limit = _load(path, sheet, first_row, limit, encoding)
    index = _column_index(column, header)
    values = []
    for row in rows[first_row - 1:]:
        value = row[index] if index < len(row) else ""
        if value:
            values.append(value)
        if len(values) >= limit:
            break
    return values
